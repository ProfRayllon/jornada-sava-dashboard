import math
import re
import unicodedata
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


RAW_2026 = Path("JORNADA SAVA 2026.xlsx")
TREATMENT_2025 = Path("Tratamneto - base 2025.xlsx")
OUTPUT = Path("Tratamento - base 2026.xlsx")

ATTENDANCE_POINTS = 575
ATTENDANCE_DATES = {
    "19/03": datetime(2026, 3, 19),
    "20/05": datetime(2026, 5, 20),
    "23/05": datetime(2026, 5, 23),
    "18/06": datetime(2026, 6, 18),
}

KNOWN_2025_TOTALS = {
    "ARTHUR MAIA SUASSUNA": 11035,
    "DAVI DE ASSIS PINHEIRO DA SILVA": 8285,
    "TAMARA MENESES MEDEIROS DE MELO": 5350,
    "BRUNO ANDRADE FARIAS": 8800,
    "MAURICIO BARBOSA DE PAIVA": 10755,
    "GABRIEL SOUTO MAIOR PEIXOTO": 7290,
    "PAULO ALVES DA SILVA FILHO": 7610,
    "JULIA RENALE OLIVEIRA": 7710,
    "CYBELLE FERNANDES DA SILVA": 8875,
    "LEONARDO ARNAUD DE LUCENA LOPES": 8450,
    "KATIA WANESSA BORGES DE LIMA LUZ": 9175,
    "RAFAEL JOSE BARRETO SERRANO": 7710,
    "GABRIEL NASCIMENTO RODRIGUES": 9195,
    "JONATAN RAULIM RAMOS": 8610,
    "DANIELY MARIA MOURA DE OLIVEIRA": 9560,
    "GILDERSON ALEXANDRE DA SILVA": 8390,
    "LORENA DE OLIVEIRA ALVES": 6760,
    "CLAYTON PEREIRA DE OLIVEIRA": 8760,
    "THAIS DANTAS CAVALCANTI": 8500,
}

LEVELS = [
    ("Explorador SAVA", 0, 6000),
    ("Protagonista SAVA", 6001, 13000),
    ("Líder SAVA", 13001, 23000),
    ("Guardião SAVA", 23001, math.inf),
]

ADV_FINAL_COLUMNS = [
    "Alta responsabilidade",
    "Relacionamentos colaborativos",
    "Conhecimentos jurídicos",
    "Proatividade estratégica",
]

ADM_FINAL_COLUMNS = [
    "Qualidade técnica nos serviços jurídicos",
    "Eficiência no atendimento ao cliente",
    "Domínio de ferramentas tecnológicas",
    "Proatividade",
    "Inteligência emocional e social",
    "Autogestão",
]

BASE_HEADERS = [
    "Nome",
    *ATTENDANCE_DATES.values(),
    "Nome do avaliador (coordenador) 4",
    "Área de atuação",
    "Período da avaliação",
    "Responsabilidade",
    "Relacionamentos colaborativos",
    "Conhecimentos jurídicos",
    "Proatividade estratégica",
    "Nome do avaliador (coordenador)",
    "Setor",
    "Data da avaliação ADM",
    "Qualidade técnica nos serviços",
    "Eficiência no atendimento ao cliente",
    "Domínio de ferramentas tecnológicas",
    "Proatividade ADM",
    "Inteligência emocional e social",
    "Autogestão",
    *ADV_FINAL_COLUMNS,
    *ADM_FINAL_COLUMNS,
    "Categoria",
    "Rank",
    "Total para nível",
    "Nível",
    "Selo Sou feliz!",
    "Selo Mente brilhante",
    "Selo Eu inspiro pessoas!",
]


def clean(value):
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except TypeError:
        pass
    return str(value).strip()


def normalize(value):
    value = clean(value).upper()
    value = "".join(
        char for char in unicodedata.normalize("NFD", value) if unicodedata.category(char) != "Mn"
    )
    value = re.sub(r"[^A-Z0-9]+", " ", value)
    return re.sub(r"\s+", " ", value).strip()


def token_set(value):
    ignored = {"DE", "DA", "DO", "DAS", "DOS", "E"}
    return {token for token in normalize(value).split() if token and token not in ignored}


def resolve_key(store, name):
    normalized = normalize(name)
    if normalized in store:
        return normalized

    tokens = token_set(name)
    if not tokens:
        return normalized

    first = next(iter(normalize(name).split()), "")
    for key, person in store.items():
        other_tokens = token_set(person["Nome"])
        other_first = next(iter(normalize(person["Nome"]).split()), "")
        if first == other_first and len(tokens & other_tokens) >= min(2, len(tokens), len(other_tokens)):
            return key
    return normalized


def to_number(value):
    if value is None:
        return 0
    try:
        if pd.isna(value):
            return 0
    except TypeError:
        pass
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0


def get_level(score):
    for name, minimum, maximum in LEVELS:
        if minimum <= score <= maximum:
            return name
    return LEVELS[0][0]


def lawyer_bright_score(legal_knowledge, strategic_proactivity):
    return (to_number(legal_knowledge) / 950 * 750) + (to_number(strategic_proactivity) / 950 * 750)


def lawyer_inspire_score(responsibility, collaboration):
    return (to_number(responsibility) / 800 * 1000) + (to_number(collaboration) / 800 * 1000)


def build_2025_level_lookup():
    if not TREATMENT_2025.exists():
        return {}

    df = pd.read_excel(TREATMENT_2025, sheet_name="base_relacionada")
    lookup = {}
    for _, row in df.iterrows():
        name = clean(row.get("Nome"))
        if not name:
            continue
        current_score = sum(to_number(row.iloc[idx]) for idx in range(23, 37))
        previous_attendance = sum(to_number(row.iloc[idx]) for idx in range(1, 7))
        previous_adv = (
            to_number(row.iloc[10]) * 160
            + to_number(row.iloc[11]) * 160
            + to_number(row.iloc[12]) * 190
            + to_number(row.iloc[13]) * 190
        )
        previous_adm = (
            to_number(row.iloc[17]) * 100
            + to_number(row.iloc[18]) * 100
            + to_number(row.iloc[19]) * 100
            + to_number(row.iloc[20]) * 120
            + to_number(row.iloc[21]) * 120
            + to_number(row.iloc[22]) * 160
        )
        calculated = round(current_score + previous_attendance + previous_adv + previous_adm)
        lookup[normalize(name)] = KNOWN_2025_TOTALS.get(normalize(name), calculated)
    return lookup


def find_level_history(history, name):
    normalized = normalize(name)
    if normalized in history:
        return history[normalized]

    tokens = token_set(name)
    first = next(iter(normalize(name).split()), "")
    for key, score in history.items():
        other_tokens = set(key.split())
        other_first = next(iter(key.split()), "")
        if first == other_first and len(tokens & other_tokens) >= min(2, len(tokens), len(other_tokens)):
            return score
    return 0


def base_record(name):
    return {header: 0 for header in BASE_HEADERS} | {
        "Nome": clean(name),
        "Categoria": "Sem avaliação",
        "Nível": "Explorador SAVA",
        "Selo Sou feliz!": "",
        "Selo Mente brilhante": "",
        "Selo Eu inspiro pessoas!": "",
    }


def build_records():
    df = pd.read_excel(RAW_2026)
    kind_col = df.columns[1]
    records = {}

    def person(name):
        key = resolve_key(records, name)
        if key not in records:
            records[key] = base_record(name)
        return records[key]

    for _, row in df.iterrows():
        kind = clean(row.get(kind_col))
        if kind == "Avaliação Advogados":
            name = clean(row.get("Nome completo do advogado (avaliado)"))
            if not name:
                continue
            rec = person(name)
            rec["Categoria"] = "Advogados Associados"
            rec["Nome do avaliador (coordenador) 4"] = clean(row.get("Nome do avaliador (coordenador)."))
            rec["Área de atuação"] = clean(row.get("Área de atuação"))
            rec["Período da avaliação"] = row.get("Data da avaliação")
            rec["Responsabilidade"] = to_number(row.get("Responsabilidade"))
            rec["Relacionamentos colaborativos"] = to_number(row.get("Relacionamentos colaborativos"))
            rec["Conhecimentos jurídicos"] = to_number(row.get("Conhecimentos jurídicos"))
            rec["Proatividade estratégica"] = to_number(row.get("Proatividade estratégica"))
            rec["Alta responsabilidade"] = rec["Responsabilidade"]
            rec["Relacionamentos colaborativos"] = rec["Relacionamentos colaborativos"]
            rec["Conhecimentos jurídicos"] = rec["Conhecimentos jurídicos"]
            rec["Proatividade estratégica"] = rec["Proatividade estratégica"]

        elif kind == "Avaliação Controladoria e ADM":
            name = clean(row.get("Nome do colaborador (avaliado)"))
            if not name:
                continue
            rec = person(name)
            rec["Categoria"] = "Controladoria e Administrativo"
            rec["Nome do avaliador (coordenador)"] = clean(row.get("Nome do avaliador (coordenador)"))
            rec["Setor"] = clean(row.get("Setor"))
            rec["Data da avaliação ADM"] = row.get("Data da avaliação 2")
            rec["Qualidade técnica nos serviços"] = to_number(row.get("Qualidade técnica nos serviços"))
            rec["Eficiência no atendimento ao cliente"] = to_number(row.get("Eficiência no atendimento ao cliente"))
            rec["Domínio de ferramentas tecnológicas"] = to_number(row.get("Domínio de ferramentas tecnológicas"))
            rec["Proatividade ADM"] = to_number(row.get("Proatividade"))
            rec["Inteligência emocional e social"] = to_number(row.get("Inteligência emocional e social"))
            rec["Autogestão"] = to_number(row.get("Autogestão"))
            rec["Qualidade técnica nos serviços jurídicos"] = rec["Qualidade técnica nos serviços"]
            rec["Eficiência no atendimento ao cliente"] = rec["Eficiência no atendimento ao cliente"]
            rec["Domínio de ferramentas tecnológicas"] = rec["Domínio de ferramentas tecnológicas"]
            rec["Proatividade"] = rec["Proatividade ADM"]
            rec["Inteligência emocional e social"] = rec["Inteligência emocional e social"]
            rec["Autogestão"] = rec["Autogestão"]

        elif kind == "Acompanhamento RH":
            for header in df.columns[24:]:
                raw_value = clean(row.get(header))
                if not raw_value or "[" not in str(header):
                    continue
                match = re.search(r"\[(.*?)\]", str(header))
                if not match:
                    continue
                rec = person(match.group(1))
                for item in [part.strip() for part in raw_value.split(",") if part.strip()]:
                    date_header = ATTENDANCE_DATES.get(item)
                    if date_header:
                        rec[date_header] = ATTENDANCE_POINTS

    history = build_2025_level_lookup()
    for rec in records.values():
        attendance = sum(to_number(rec[date_header]) for date_header in ATTENDANCE_DATES.values())
        adv_score = sum(to_number(rec[col]) for col in ADV_FINAL_COLUMNS)
        adm_score = sum(to_number(rec[col]) for col in ADM_FINAL_COLUMNS)
        rank = round(attendance + adv_score + adm_score)
        previous = find_level_history(history, rec["Nome"])
        rec["Rank"] = rank
        rec["Total para nível"] = rank + previous
        rec["Nível"] = get_level(rec["Total para nível"])

        if attendance >= ATTENDANCE_POINTS * len(ATTENDANCE_DATES):
            rec["Selo Sou feliz!"] = "Sim"
        if (
            rec["Categoria"] == "Advogados Associados"
            and lawyer_bright_score(rec["Conhecimentos jurídicos"], rec["Proatividade estratégica"]) >= 1500
        ) or (
            rec["Categoria"] == "Controladoria e Administrativo"
            and sum(to_number(rec[col]) for col in ADM_FINAL_COLUMNS[:3]) >= 1500
        ):
            rec["Selo Mente brilhante"] = "Sim"
        if (
            rec["Categoria"] == "Advogados Associados"
            and lawyer_inspire_score(rec["Alta responsabilidade"], rec["Relacionamentos colaborativos"]) >= 2000
        ) or (
            rec["Categoria"] == "Controladoria e Administrativo"
            and sum(to_number(rec[col]) for col in ADM_FINAL_COLUMNS[3:]) >= 2000
        ):
            rec["Selo Eu inspiro pessoas!"] = "Sim"

    return sorted(records.values(), key=lambda item: (-item["Rank"], item["Nome"]))


def write_workbook(records):
    wb = Workbook()
    ws = wb.active
    ws.title = "base_relacionada"
    ws.append(BASE_HEADERS)
    for rec in records:
        ws.append([rec.get(header, 0) for header in BASE_HEADERS])

    raw_ws = wb.create_sheet("raw_2026")
    raw_df = pd.read_excel(RAW_2026)
    raw_ws.append(list(raw_df.columns))
    for row in raw_df.itertuples(index=False, name=None):
        raw_ws.append(list(row))

    summary_ws = wb.create_sheet("resumo")
    summary_ws.append(["Indicador", "Valor"])
    summary_ws.append(["Pessoas", len(records)])
    summary_ws.append(["Pontuação máxima do ano", max((record["Rank"] for record in records), default=0)])
    summary_ws.append(["Total Programa de Felicidade", sum(record[date] for record in records for date in ATTENDANCE_DATES.values())])
    summary_ws.append(["Total Eventos institucionais", 0])
    summary_ws.append(["Total Avaliação técnica", sum(sum(record[col] for col in ADM_FINAL_COLUMNS[:3]) for record in records)])
    summary_ws.append(["Total Avaliação socioemocional", sum(sum(record[col] for col in ADM_FINAL_COLUMNS[3:]) for record in records)])
    summary_ws.append(["Total Acompanhamento de competências", sum(sum(record[col] for col in ADV_FINAL_COLUMNS) for record in records)])

    style_workbook(wb)
    try:
        wb.save(OUTPUT)
    except PermissionError:
        fallback = OUTPUT.with_name(f"{OUTPUT.stem} - atualizado.xlsx")
        wb.save(fallback)
        print(f"{OUTPUT} está aberto ou bloqueado. Cópia atualizada salva em {fallback}")


def style_workbook(wb):
    header_fill = PatternFill("solid", fgColor="87800F")
    header_font = Font(color="FFFFFF", bold=True)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col_idx, column_cells in enumerate(ws.columns, start=1):
            max_len = 8
            for cell in column_cells:
                value = cell.value
                if isinstance(value, datetime):
                    cell.number_format = "dd/mm/yyyy"
                    max_len = max(max_len, 12)
                else:
                    max_len = max(max_len, min(len(clean(value)), 42))
                if isinstance(value, (int, float)):
                    cell.number_format = '#,##0'
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 34)
        ws.auto_filter.ref = ws.dimensions


def main():
    records = build_records()
    write_workbook(records)
    print(f"{OUTPUT} criada com {len(records)} pessoas")
    for rec in records[:8]:
        print(rec["Nome"], rec["Rank"], rec["Total para nível"], rec["Nível"])


if __name__ == "__main__":
    main()
